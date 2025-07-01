import pandas as pd
import urllib.parse
import re
from datetime import datetime
from rich.console import Console
from rich.progress import track
from rich.panel import Panel
from rich.text import Text
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from collections import defaultdict
from pyfiglet import Figlet

# === Setup console ===
console = Console()
attack_count = defaultdict(int)
detected_rows = []
total_scanned = 0

# === Cyberpunk-style banner ===
def show_banner():
    f = Figlet(font='slant')
    banner = f.renderText('Malicious URI Detector')
    banner_text = Text(banner, style="bold magenta")
    console.print(banner_text)

    author_text = Text("Author: Yogesh", style="bold white on rgb(18,18,18)")
    cyber_panel = Panel(
        author_text,
        title="🧠 INFO",
        title_align="left",
        border_style="bright_cyan",
        style="on rgb(30,0,45)"
    )
    console.print(cyber_panel)

# === Enhanced Regex Signatures ===
attack_patterns = {
    # SQL Injection - Enhanced with more variants and obfuscation techniques
    "SQL Injection": re.compile(r"(?i)(\b(UNION\s+(ALL\s+)?SELECT|SELECT\s+.+\s+FROM|INSERT\s+INTO|UPDATE\s+\w+\s+SET|DELETE\s+FROM|DROP\s+(TABLE|DATABASE)|ALTER\s+TABLE|CREATE\s+(TABLE|DATABASE))\b|"
                               r"\b(SLEEP|BENCHMARK|WAITFOR\s+DELAY|PG_SLEEP)\s*\(|"
                               r"(--|#|\/\*|\*\/|;\s*--|;\s*\/\*)|"
                               r"(\'\s*(OR|AND)\s*\'|\'\s*(OR|AND)\s*\d+\s*=\s*\d+|1\s*=\s*1|0\s*=\s*0)|"
                               r"(\bCHAR\s*\(|\bCONCAT\s*\(|\bASCII\s*\(|\bSUBSTRING\s*\(|\bVERSION\s*\(\))|"
                               r"(0x[0-9a-fA-F]+|CHAR\([0-9,\s]+\))|"
                               r"(\'\s*(;|--|#)|\'\s*\+\s*\'|\'\s*\|\|\s*\')|"
                               r"(\bINFORMATION_SCHEMA\b|\bmysql\.user\b|\bsysdatabases\b|\bsysusers\b)|"
                               r"(\bLOAD_FILE\s*\(|\bINTO\s+OUTFILE\b|\bINTO\s+DUMPFILE\b))"),

    # Command Injection - More comprehensive OS command detection
    "Command Injection": re.compile(r"(?i)([\s;&|`$]\s*(ls|cat|grep|awk|sed|find|whoami|id|pwd|uname|ps|netstat|ifconfig|curl|wget|nc|ncat|telnet|ssh|ftp|ping|nslookup|dig)\b|"
                                   r"[\s;&|]\s*(bash|sh|zsh|csh|tcsh|cmd|powershell|wmic)\b|"
                                   r"`[^`]*`|\$\([^)]*\)|\$\{[^}]*\}|"
                                   r"(\|\s*(cat|more|less|head|tail)\b)|"
                                   r"(>\s*/[a-zA-Z0-9_/.-]+|>>\s*/[a-zA-Z0-9_/.-]+)|"
                                   r"(&&\s*[a-zA-Z_][a-zA-Z0-9_]*|;\s*[a-zA-Z_][a-zA-Z0-9_]*))"),

    # XSS - Enhanced with more vectors and encoding variations
    "XSS (Script-based)": re.compile(r"(?i)(<\s*script[^>]*>.*?<\s*/\s*script\s*>|"
                                     r"<\s*script[^>]*>|"
                                     r"javascript\s*:|"
                                     r"on(load|click|error|focus|blur|change|submit|mouseover|mouseout|keyup|keydown)\s*=|"
                                     r"(alert|confirm|prompt|eval|setTimeout|setInterval)\s*\(|"
                                     r"document\.(cookie|write|writeln|location|domain)|"
                                     r"window\.(location|open)|"
                                     r"String\.fromCharCode\s*\(|"
                                     r"unescape\s*\(|"
                                     r"<\s*(img|iframe|object|embed|link|meta)[^>]*(src|href)\s*=\s*[\"']?\s*javascript:|"
                                     r"<\s*svg[^>]*onload\s*=)"),

    # XSS (Event Handler Injection)
    "XSS (Event Handlers)": re.compile(r"(?i)(<[^>]+\s+(on\w+)\s*=\s*[\"']?[^\"'>]*[\"']?[^>]*>|"
                                       r"\s+(on\w+)\s*=\s*[\"']?[^\"'>]*[\"']?|"
                                       r"<\s*(input|button|a|div|span|img)[^>]*\s+on\w+\s*=)"),

    # XSS (DOM-based)
    "XSS (DOM-based)": re.compile(r"(?i)(document\.(location|URL|referrer|write|writeln)|"
                                  r"window\.(location|name)|"
                                  r"history\.(pushState|replaceState)|"
                                  r"innerHTML\s*=|"
                                  r"outerHTML\s*=|"
                                  r"insertAdjacentHTML\s*\(|"
                                  r"\.src\s*=\s*[\"']?javascript:|"
                                  r"\.href\s*=\s*[\"']?javascript:)"),

    # HTML Injection - More specific patterns
    "HTML Injection": re.compile(r"(?i)(<\s*(iframe|object|embed|applet|form|input|textarea|select|button|meta|link)\s*[^>]*>|"
                                 r"<\s*(h[1-6]|div|span|p|a|img)\s+[^>]*style\s*=\s*[\"'][^\"'>]*[\"']|"
                                 r"<\s*style[^>]*>.*?<\s*/\s*style\s*>|"
                                 r"<\s*link[^>]+rel\s*=\s*[\"']?stylesheet[\"']?|"
                                 r"<\s*base\s+href\s*=)"),

    # Path Traversal - Enhanced with various encoding and techniques
    "Path Traversal": re.compile(r"(?i)((\.\./|\.\.\\|%2e%2e%2f|%2e%2e\\|\.\.%2f|\.\.%5c){2,}|"
                                 r"(\.\.%252f|\.\.%255c|%252e%252e%252f|%252e%252e%255c)|"
                                 r"(\.\./.*/(etc/passwd|boot\.ini|windows/system32|proc/self/environ)|"
                                 r"\\\.\.\\.*\\(windows\\system32|boot\.ini))|"
                                 r"(\.\./){3,}|"
                                 r"(file://|file:///))"),

    # LDAP Injection - More comprehensive
    "LDAP Injection": re.compile(r"(?i)((\*|\(|\)|&|\||!)\s*(\)|&|\||=|>|<)|"
                                 r"(uid\s*=\s*\*|cn\s*=\s*\*|ou\s*=\s*\*|dc\s*=\s*\*)|"
                                 r"(\(\s*\|\s*\(|\(\s*&\s*\()|"
                                 r"(objectClass\s*=\s*\*|sAMAccountName\s*=)|"
                                 r"(admin|administrator|root|guest)\*|"
                                 r"(\)\s*\(\s*\||\)\s*\(\s*&))"),

    # Command Injection (Windows specific)
    "Command Injection (Windows)": re.compile(r"(?i)(cmd\.exe|powershell\.exe|wmic\.exe|net\.exe|"
                                              r"dir\s+/|type\s+\w+|copy\s+\w+|move\s+\w+|del\s+\w+|"
                                              r"echo\s+.*>\s*\w+|"
                                              r"&\s*(dir|type|copy|del|net|wmic)|"
                                              r"\|\s*(findstr|find|more)|"
                                              r"%[A-Za-z_][A-Za-z0-9_]*%|"
                                              r"\$env:[A-Za-z_][A-Za-z0-9_]*|"
                                              r"Get-Process|Get-Service|Invoke-Expression)"),

    # SSRF - Enhanced with more targets and bypass techniques
    "SSRF": re.compile(r"(?i)((https?://)?(localhost|127\.0\.0\.1|0\.0\.0\.0|::1|0x7f000001|2130706433)|"
                      r"(https?://)?10\.(25[0-5]|2[0-4][0-9]|1[0-9][0-9]|[1-9]?[0-9])\.(25[0-5]|2[0-4][0-9]|1[0-9][0-9]|[1-9]?[0-9])\.(25[0-5]|2[0-4][0-9]|1[0-9][0-9]|[1-9]?[0-9])|"
                      r"(https?://)?192\.168\.(25[0-5]|2[0-4][0-9]|1[0-9][0-9]|[1-9]?[0-9])\.(25[0-5]|2[0-4][0-9]|1[0-9][0-9]|[1-9]?[0-9])|"
                      r"(https?://)?172\.(1[6-9]|2[0-9]|3[01])\.(25[0-5]|2[0-4][0-9]|1[0-9][0-9]|[1-9]?[0-9])\.(25[0-5]|2[0-4][0-9]|1[0-9][0-9]|[1-9]?[0-9])|"
                      r"(https?://)?169\.254\.(25[0-5]|2[0-4][0-9]|1[0-9][0-9]|[1-9]?[0-9])\.(25[0-5]|2[0-4][0-9]|1[0-9][0-9]|[1-9]?[0-9])|"
                      r"file://|gopher://|dict://|ftp://internal|ldap://|sftp://)"),

    # RCE - Enhanced with more functions and techniques
    "RCE": re.compile(r"(?i)(\b(eval|exec|system|shell_exec|passthru|popen|proc_open|assert|call_user_func|create_function|preg_replace.*\/e|file_get_contents|file_put_contents|fopen|fwrite|include|require|include_once|require_once)\s*\(|"
                     r"\b(Runtime\.getRuntime|ProcessBuilder|eval|Function|setTimeout|setInterval)\s*\(|"
                     r"`[^`]*`|\$\([^)]*\)|"
                     r"\b(base64_decode|hex2bin|gzinflate|str_rot13|strrev)\s*\(.*\b(eval|exec|system|shell_exec)\b|"
                     r"\\x[0-9a-fA-F]{2}|"
                     r"chr\s*\(\s*\d+\s*\)|"
                     r"__import__\s*\(|"
                     r"getattr\s*\(.*,\s*[\"'].*[\"']\s*\))"),

    # XXE Injection - More comprehensive
    "XXE": re.compile(r"(?i)(<!DOCTYPE\s+[^>]*\[.*?<!ENTITY\s+.*?SYSTEM\s+[\"']?(file://|http://|ftp://)|"
                     r"<!ENTITY\s+.*?PUBLIC\s+.*?SYSTEM|"
                     r"<!ENTITY\s+.*?%\s+.*?SYSTEM|"
                     r"ENTITY\s+.*?file://|"
                     r"ENTITY\s+.*?/etc/passwd|"
                     r"ENTITY\s+.*?/proc/self/environ|"
                     r"<!ENTITY.*?>.*?&.*?;)"),

    # CRLF Injection - Enhanced
    "CRLF Injection": re.compile(r"(?i)((%0d%0a|%0a%0d|%0a|%0d|\\r\\n|\\n\\r|\\r|\\n|\r\n|\n\r|\r|\n)\s*(Set-Cookie:|Location:|Content-Type:|Content-Length:|X-.*:|Cache-Control:|Expires:|Last-Modified:)|"
                                 r"(%0d%0a|%0a|%0d|\\r\\n|\\n|\\r|\r\n|\n|\r)(%0d%0a|%0a|%0d|\\r\\n|\\n|\\r|\r\n|\n|\r))"),

    # Open Redirect - Enhanced with more parameters and encoding
    "Open Redirect": re.compile(r"(?i)((\?|&)(redirect|url|next|target|return|returnurl|goto|link|redir|dest|destination|continue|success|forward|r|u|to|ref|referer|site|domain|host|page|view|redirect_uri|return_to|callback|exit|out|jump)\s*=\s*(https?://[^&\s]+|//[^&\s]+|[^&\s]*\.\.)|"
                               r"window\.location\s*=\s*[\"']?https?://|"
                               r"document\.location\s*=\s*[\"']?https?://|"
                               r"location\.href\s*=\s*[\"']?https?://)"),

    # JWT Manipulation - Enhanced
    "JWT Manipulation": re.compile(r"(?i)(eyJ[a-zA-Z0-9_-]*\.eyJ[a-zA-Z0-9_-]*\.(eyJ[a-zA-Z0-9_-]*)?|"
                                   r"Bearer\s+eyJ[a-zA-Z0-9_-]*\.|"
                                   r"authorization:\s*bearer\s+eyJ|"
                                   r"jwt\s*=\s*eyJ|"
                                   r"token\s*=\s*eyJ)"),

    # Template Injection - Enhanced with more engines
    "Template Injection": re.compile(r"(?i)(\{\{.*?\}\}|\{%.*?%\}|\{\{.*?\|\s*(safe|escape|raw)\}\}|"
                                     r"\$\{.*?\}|<%.*?%>|<\?.*?\?>|"
                                     r"\{\{.*?(config|request|session|g|url_for|get_flashed_messages).*?\}\}|"
                                     r"\{\{.*?__.*__.*?\}\}|"
                                     r"\{\{.*?(class|mro|subclasses|globals|builtins).*?\}\}|"
                                     r"<#.*?#>|\[@.*?@\]|#\{.*?\})"),

    # Prototype Pollution - Enhanced
    "Prototype Pollution": re.compile(r"(?i)(__proto__|constructor\.prototype|prototype\.constructor|"
                                      r"\[\"__proto__\"\]|\['__proto__'\]|"
                                      r"\.constructor\.prototype\.|"
                                      r"\[\"constructor\"\]\[\"prototype\"\]|"
                                      r"Object\.prototype|"
                                      r"constructor\[\"prototype\"\])"),

    # NoSQL Injection - New addition
    "NoSQL Injection": re.compile(r"(?i)(\$where|\$ne|\$gt|\$lt|\$gte|\$lte|\$in|\$nin|\$regex|\$exists|\$type|\$size|\$all|\$elemMatch|"
                                  r"\[\$ne\]|\[\$gt\]|\[\$lt\]|\[\$gte\]|\[\$lte\]|\[\$in\]|\[\$nin\]|\[\$regex\]|\[\$where\]|"
                                  r"\.find\s*\(.*\$|\.update\s*\(.*\$|\.remove\s*\(.*\$|"
                                  r"true.*==.*true|false.*!=.*false|"
                                  r"\|\|.*==|&&.*!=|sleep\s*\(\d+\))"),

    # File Upload - New addition
    "File Upload Bypass": re.compile(r"(?i)(\.php\.|\.asp\.|\.jsp\.|\.py\.|\.rb\.|\.pl\.|"
                                     r"\.php[3-7]?$|\.asp$|\.aspx$|\.jsp$|\.jspx$|\.py$|\.rb$|\.pl$|\.cgi$|"
                                     r"\.htaccess|\.htpasswd|web\.config|"
                                     r"null\.php|file\.php|upload\.php|shell\.php|cmd\.php|"
                                     r"Content-Type:\s*image/.*\.php|"
                                     r"filename=.*\.php|filename=.*\.asp|filename=.*\.jsp)"),

    # XML Injection - New addition
    "XML Injection": re.compile(r"(?i)(<\?xml[^>]*>.*?<.*?>|"
                               r"<!DOCTYPE[^>]*>|"
                               r"<!\[CDATA\[.*?\]\]>|"
                               r"&#x[0-9a-fA-F]+;|&#\d+;|"
                               r"<[^>]*>[^<]*<[^>]*script[^>]*>|"
                               r"<[^>]*\s+(xmlns|xsi:|xsd:)[^>]*>)"),

    # HTTP Parameter Pollution - Enhanced
    "HTTP Parameter Pollution": re.compile(r"(?i)((\?|&)[^=&]*=([^&]*&){2,}|"
                                          r"(\?|&)[^=&]*=[^&]*&[^=&]*=[^&]*&[^=&]*=|"
                                          r"(\?|&)(id|user|admin|role|auth|token)=[^&]*&(id|user|admin|role|auth|token)=)"),

    # Email Header Injection - Enhanced
    "Email Header Injection": re.compile(r"(?i)((%0d%0a|%0a|%0d|\\r\\n|\\n|\\r|\r\n|\n|\r)\s*(to|cc|bcc|from|subject|reply-to|return-path|x-mailer|message-id|date|mime-version|content-type|content-transfer-encoding):|"
                                         r"(to|cc|bcc):\s*[^@\s]+@[^@\s]+\s*(,|;)\s*[^@\s]+@|"
                                         r"subject:\s*.*?(%0d%0a|\\r\\n|\r\n)|"
                                         r"x-mailer:\s*|content-type:\s*text/html)"),

    # LDAP Injection - Additional patterns
    "LDAP Injection Advanced": re.compile(r"(?i)(\)\s*\(\s*\|\s*\(.*?=.*?\*|"
                                         r"\)\s*\(\s*&\s*\(.*?=.*?\*|"
                                         r"\*\s*\)\s*\(\s*objectclass\s*=|"
                                         r"\(\s*cn\s*=\s*admin\s*\)|"
                                         r"\(\s*uid\s*=\s*0\s*\)|"
                                         r"userPassword\s*=\s*\*|"
                                         r"memberOf\s*=\s*\*)"),

    # File Inclusion - Local and Remote
    "File Inclusion": re.compile(r"(?i)((https?|ftp)://[^/]+/.*\.(php|asp|jsp|py|rb|pl|txt|log|conf|ini|cfg)|"
                                r"(\.\./)*(etc/passwd|proc/self/environ|var/log|windows/system32|boot\.ini)|"
                                r"(include|require|include_once|require_once)\s*\(\s*[\"']?(https?://|\.\./).*?\.|"
                                r"php://input|php://filter|data://|expect://|zip://|"
                                r"file:///(etc/passwd|windows/system32|boot\.ini))"),
}

# === Utility Functions ===
def decode_uri(uri):
    """Decode URI with multiple rounds to catch double encoding"""
    try:
        decoded = uri
        # Try multiple rounds of decoding to catch double/triple encoding
        for _ in range(3):
            prev_decoded = decoded
            decoded = urllib.parse.unquote(decoded)
            if decoded == prev_decoded:
                break
        return decoded
    except:
        return uri

def normalize_uri(uri):
    """Normalize URI for better detection"""
    try:
        # Convert to lowercase for case-insensitive matching
        normalized = uri.lower()
        # Replace common URL encoding variations
        normalized = normalized.replace('%20', ' ')
        normalized = normalized.replace('+', ' ')
        return normalized
    except:
        return uri

# === Scan Logic ===
def scan_excel_log(file_path):
    global total_scanned
    
    try:
        df = pd.read_excel(file_path, engine='openpyxl')
    except FileNotFoundError:
        console.print(f"[bold red]❌ File '{file_path}' not found. Please check the file path.[/bold red]")
        return
    except Exception as e:
        console.print(f"[bold red]❌ Error reading Excel file: {str(e)}[/bold red]")
        return

    if "requestUri_s" not in df.columns:
        console.print("[bold red]❌ Column 'requestUri_s' not found in Excel file.[/bold red]")
        console.print(f"[yellow]Available columns: {', '.join(df.columns)}[/yellow]")
        return

    console.print("[bold cyan]🔍 Scanning URIs for attack signatures...[/bold cyan]")
    
    for row in track(df.itertuples(index=False), description="Analyzing..."):
        original_uri = str(getattr(row, 'requestUri_s'))
        decoded_uri = decode_uri(original_uri)
        normalized_uri = normalize_uri(decoded_uri)
        
        total_scanned += 1
        matched_attacks = []

        # Check against all patterns
        for attack, pattern in attack_patterns.items():
            # Check both decoded and normalized versions
            if pattern.search(decoded_uri) or pattern.search(normalized_uri):
                matched_attacks.append(attack)
                attack_count[attack] += 1

        if matched_attacks:
            detected_rows.append({
                "Request URI": original_uri[:500] + "..." if len(original_uri) > 500 else original_uri,  # Truncate very long URIs
                "Decoded URI": decoded_uri[:500] + "..." if len(decoded_uri) > 500 else decoded_uri,
                "Detected Attack Type(s)": ", ".join(matched_attacks),
                "Severity": get_severity(matched_attacks)
            })

    generate_excel_report()

def get_severity(attacks):
    """Assign severity based on attack types"""
    high_severity = ["RCE", "Command Injection", "SQL Injection", "XXE", "File Inclusion"]
    medium_severity = ["XSS", "SSRF", "Template Injection", "Path Traversal"]
    
    for attack in attacks:
        for high in high_severity:
            if high in attack:
                return "HIGH"
        for medium in medium_severity:
            if medium in attack:
                return "MEDIUM"
    return "LOW"

# === Excel Report Generation ===
def generate_excel_report():
    if not detected_rows:
        console.print("[bold green]✅ No malicious URIs found. All clean.[/bold green]")
        print_summary()
        return

    timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"malicious_uri_scan_report_{timestamp_str}.xlsx"
    df_report = pd.DataFrame(detected_rows)
    df_report.to_excel(output_file, index=False)

    # Style the Excel file
    wb = load_workbook(output_file)
    ws = wb.active

    # Header formatting
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Severity color coding
    high_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    medium_fill = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")
    low_fill = PatternFill(start_color="A8E6CF", end_color="A8E6CF", fill_type="solid")

    # Format headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Format severity column based on values
    severity_col = None
    for col_num, cell in enumerate(ws[1], 1):
        if cell.value == "Severity":
            severity_col = col_num
            break

    if severity_col:
        for row in ws.iter_rows(min_row=2, min_col=severity_col, max_col=severity_col):
            for cell in row:
                if cell.value == "HIGH":
                    cell.fill = high_fill
                elif cell.value == "MEDIUM":
                    cell.fill = medium_fill
                elif cell.value == "LOW":
                    cell.fill = low_fill

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 100)  # Cap at 100 characters
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_file)
    console.print(f"[bold cyan]📄 Report saved as: [underline]{output_file}[/underline][/bold cyan]")
    print_summary()

# === Summary Table ===
def print_summary():
    from rich.table import Table
    table = Table(title="🧾 Attack Summary", header_style="bold green")
    table.add_column("Attack Type", style="cyan")
    table.add_column("Occurrences", style="yellow", justify="right")
    table.add_column("Risk Level", style="red", justify="center")

    total_detected = sum(attack_count.values())
    
    # Sort by count (descending)
    sorted_attacks = sorted(attack_count.items(), key=lambda x: x[1], reverse=True)
    
    for attack, count in sorted_attacks:
        severity = get_severity([attack])
        risk_color = "red" if severity == "HIGH" else "yellow" if severity == "MEDIUM" else "green"
        table.add_row(attack, str(count), f"[{risk_color}]{severity}[/{risk_color}]")

    table.add_row("[bold white]Total Attacks Detected[/bold white]", f"[bold yellow]{total_detected}[/bold yellow]", "")
    table.add_row("[bold white]Total URIs Scanned[/bold white]", f"[bold cyan]{total_scanned}[/bold cyan]", "")
    
    if total_scanned > 0:
        detection_rate = (total_detected / total_scanned) * 100
        table.add_row("[bold white]Detection Rate[/bold white]", f"[bold magenta]{detection_rate:.2f}%[/bold magenta]", "")
    
    console.print(table)

# === Main Execution ===
if __name__ == "__main__":
    show_banner()
    
    # Check if logs.xlsx exists, otherwise prompt for file
    import os
    log_file = "logs.xlsx"
    
    if not os.path.exists(log_file):
        console.print(f"[yellow]⚠️  '{log_file}' not found in current directory.[/yellow]")
        console.print("[cyan]Please ensure your Excel log file is named 'logs.xlsx' or modify the filename in the script.[/cyan]")
    else:
        scan_excel_log(log_file)